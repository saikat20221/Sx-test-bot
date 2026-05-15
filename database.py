"""
database.py  —  JSON-file backend (no PostgreSQL required)
=============================================================
Architecture for 15 000 + users:
  * All data loaded into RAM at startup (instant reads, zero SQL round-trips)
  * A single RLock guards every mutation (thread-safe)
  * A background daemon thread flushes dirty data to disk every 5 s
  * Atomic writes: write -> temp file, then os.replace() (never corrupts)
  * Two data files:
        data/numbers.json  — countries, numbers, assignments, country OTP bonus
        data/users.json    — users, admins, settings, panels, panel_status,
                             referral_log, withdraw_requests, otp_bonus_log,
                             otp_deliveries, website_messages
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import os
import re
import threading
import time
import zipfile
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from config import PROTECTED_ADMINS

logger = logging.getLogger(__name__)

# ── File paths ─────────────────────────────────────────────────────────────────
_BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR     = os.path.join(_BASE_DIR, "data")
_NUMBERS_FILE = os.path.join(_DATA_DIR, "numbers.json")
_USERS_FILE   = os.path.join(_DATA_DIR, "users.json")

# ── In-memory stores ───────────────────────────────────────────────────────────
_N: dict[str, Any] = {}
_U: dict[str, Any] = {}

_lock           = threading.RLock()
_n_dirty        = False
_u_dirty        = False
_FLUSH_INTERVAL = 5   # seconds between disk flushes

# ── Default skeletons ──────────────────────────────────────────────────────────
_DEFAULT_NUMBERS: dict = {
    "next_country_id":    1,
    "next_number_id":     1,
    "next_assignment_id": 1,
    "countries":          {},
    "numbers":            {},
    "assignments":        {},
    "country_otp_bonus":  {},
}

_DEFAULT_USERS: dict = {
    "next_user_row_id":    1,
    "next_admin_id":       1,
    "next_ref_id":         1,
    "next_withdraw_id":    1,
    "next_otp_bonus_id":   1,
    "next_delivery_id":    1,
    "next_website_msg_id": 1,
    "next_panel_id":       1,
    "users":               {},
    "admins":              {},
    "referral_log":        {},
    "withdraw_requests":   {},
    "otp_bonus_log":       {},
    "otp_deliveries":      {},
    "website_messages":    {},
    "settings": {
        "referral_enabled":      "1",
        "referral_bonus":        "10.0",
        "min_withdraw":          "50.0",
        "otp_bonus_enabled":     "1",
        "otp_bonus_amount":      "2.0",
        "otp_bonus_daily_limit": "10",
        "number_limit":          "0",
        "bot_link_sxsmsbot":     "https://t.me/SXSMSBOT",
        "bot_link_unofficial":   "https://t.me/UnofficialNumberBot",
        "bot_link_channel":      "https://t.me/sxchannel143",
        "bot_link_support":      "https://t.me/sxsupportgroup",
        "channel_check_interval": "60",
        "global_services":        [],
    },
    "panels":            {},
    "panel_status":      {},
    "extra_groups":      [],
    "group_members":     {},
    "required_channels": [],
}

_DEFAULT_PANELS = [
    ("SMS Hadi",        "saikat2007",   "saikat2007",    "http://smshadi.net"),
    ("Konekta Premium", "MDSaikat",     "@saikat2007",   "https://konektapremium.net"),
    ("Msi sms",         "saikat",       "112233",        "http://145.239.130.45/ints"),
    ("Number Panel",    "forid579",     "1ssaikatyt",    "http://51.89.99.105/NumberPanel"),
    ("Purple sms",      "sxunofficial", "sxunofficial",  "http://85.195.94.50/sms"),
    ("Proof sms",       "saikatff143",  "saikatff143",   "http://217.182.195.194/ints"),
    ("Lamix sms",       "saikatff143",  "saikatff143",   "http://51.210.208.26/ints"),
    ("Seven 1 Tel",     "MDSaikat",     "saikatff143",   "http://94.23.120.156/ints"),
    ("Flex sms",        "saikatf143",   "Saikatf@12",    "http://168.119.13.175/ints"),
    ("Zento sms",       "saikatff143",  "saikatff143",   "http://54.38.176.48/ints"),
    ("Wolf sms",        "saikatff143",  "saikatff143",   "http://213.32.24.208/ints"),
    ("Shark sms",       "saikatff143",  "saikatff143",   "http://65.109.111.158/ints"),
    ("SMS Hadi 2",      "saikatff143",  "saikatff143",   "http://smshadi.net"),
]

# ── Admin cache ────────────────────────────────────────────────────────────────
_admin_cache: set | None = None
_admin_cache_ts: float   = 0.0
_ADMIN_CACHE_TTL          = 300.0


def _invalidate_admin_cache():
    global _admin_cache, _admin_cache_ts
    _admin_cache    = None
    _admin_cache_ts = 0.0


def _get_cached_admins() -> set:
    global _admin_cache, _admin_cache_ts
    now = time.monotonic()
    if _admin_cache is not None and (now - _admin_cache_ts) < _ADMIN_CACHE_TTL:
        return _admin_cache
    with _lock:
        _admin_cache    = set(_U.get("admins", {}).keys())
        _admin_cache_ts = now
    return _admin_cache


# ── Internal helpers ───────────────────────────────────────────────────────────

def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _today() -> str:
    return date.today().strftime("%Y-%m-%d")

def _mark_n_dirty():
    global _n_dirty
    _n_dirty = True

def _mark_u_dirty():
    global _u_dirty
    _u_dirty = True

def _atomic_write(path: str, data: dict):
    tmp = path + ".tmp"
    os.makedirs(os.path.dirname(tmp), exist_ok=True)
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=None, separators=(",", ":"))
    os.replace(tmp, path)

def _flush():
    global _n_dirty, _u_dirty
    if _n_dirty:
        try:
            _atomic_write(_NUMBERS_FILE, _N)
            _n_dirty = False
        except Exception as exc:
            logger.error(f"[DB] flush numbers.json failed: {exc}")
    if _u_dirty:
        try:
            _atomic_write(_USERS_FILE, _U)
            _u_dirty = False
        except Exception as exc:
            logger.error(f"[DB] flush users.json failed: {exc}")

def _flush_loop():
    while True:
        time.sleep(_FLUSH_INTERVAL)
        _flush()


# ── Old-record cleanup (runs every hour, keeps last 7 days) ───────────────────
_RETENTION_SECONDS = 30 * 24 * 3600     # 7 days
_CLEANUP_INTERVAL  = 3600              # 1 hour


def _parse_ts(ts: str) -> float:
    try:
        return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").timestamp()
    except Exception:
        return 0.0


def _cleanup_old_records() -> tuple[int, int]:
    """Delete website_messages and otp_deliveries older than 7 days."""
    cutoff = time.time() - _RETENTION_SECONDS
    removed_msgs = removed_deliv = 0
    with _lock:
        # website_messages
        wm = _U.get("website_messages", {})
        old_keys = [k for k, r in wm.items()
                    if _parse_ts(r.get("received_at", "")) < cutoff]
        for k in old_keys:
            del wm[k]
        removed_msgs = len(old_keys)

        # otp_deliveries (value is timestamp string)
        od = _U.get("otp_deliveries", {})
        old_keys = [k for k, ts in od.items() if _parse_ts(ts) < cutoff]
        for k in old_keys:
            del od[k]
        removed_deliv = len(old_keys)

        if removed_msgs or removed_deliv:
            # Reset latest caches in case the cached record was removed
            _latest_msg_cache["record"] = None
            _panel_latest_cache.clear()
            _mark_u_dirty()
    return removed_msgs, removed_deliv


def _cleanup_loop():
    while True:
        time.sleep(_CLEANUP_INTERVAL)
        try:
            m, d = _cleanup_old_records()
            if m or d:
                logger.info(f"[DB] Cleanup: removed {m} website_messages, {d} otp_deliveries (>7d)")
        except Exception as exc:
            logger.error(f"[DB] cleanup failed: {exc}")

def _generate_referral_code(user_id) -> str:
    return hashlib.md5(f"ref_{user_id}_sx".encode()).hexdigest()[:8].upper()


# ── Init ───────────────────────────────────────────────────────────────────────

def _init_db():
    global _N, _U
    os.makedirs(_DATA_DIR, exist_ok=True)

    if os.path.exists(_NUMBERS_FILE):
        try:
            with open(_NUMBERS_FILE, "r", encoding="utf-8") as f:
                _N = json.load(f)
        except Exception:
            _N = {}
    for k, v in _DEFAULT_NUMBERS.items():
        _N.setdefault(k, v)

    if os.path.exists(_USERS_FILE):
        try:
            with open(_USERS_FILE, "r", encoding="utf-8") as f:
                _U = json.load(f)
        except Exception:
            _U = {}
    for k, v in _DEFAULT_USERS.items():
        _U.setdefault(k, v)

    for k, v in _DEFAULT_USERS["settings"].items():
        _U["settings"].setdefault(k, v)

    for name in PROTECTED_ADMINS:
        if name not in _U["admins"]:
            _U["admins"][name] = {"username": name, "added_at": _now()}
            _mark_u_dirty()

    for name, uname, pwd, url in _DEFAULT_PANELS:
        if name not in _U["panels"]:
            pid = _U["next_panel_id"]
            _U["panels"][name] = {
                "id": pid, "name": name,
                "username": uname, "password": pwd, "base_url": url,
            }
            _U["next_panel_id"] = pid + 1
            _mark_u_dirty()

    _PANEL_MIGRATIONS = [
        # (panel_name, old_url,                          new_url,                       new_user,      new_pass)
        ('SMS Hadi',  'http://185.2.83.39/ints',        'http://smshadi.net',           'saikat2007',  'saikat2007'),
        ('SMS Hadi',  'http://2.59.169.96/ints',        'http://smshadi.net',           'saikat2007',  'saikat2007'),
        ('Lamix sms', 'http://139.99.208.63/ints',      'http://51.210.208.26/ints',    'saikatff143', 'saikatff143'),
    ]
    for pname, old_url, new_url, new_user, new_pass in _PANEL_MIGRATIONS:
        p = _U["panels"].get(pname)
        if p and p.get('base_url', '').rstrip('/') == old_url.rstrip('/'):
            p['base_url']  = new_url
            p['username']  = new_user
            p['password']  = new_pass
            _mark_u_dirty()
            logger.info(f"[DB] Migrated {pname}: updated URL and credentials.")

    # Sync credentials from _DEFAULT_PANELS into the live DB (runs every startup)
    _default_creds = {name: (uname, pwd) for name, uname, pwd, _url in _DEFAULT_PANELS}
    for pname, (new_user, new_pass) in _default_creds.items():
        p = _U["panels"].get(pname)
        if p and (p.get('username') != new_user or p.get('password') != new_pass):
            p['username'] = new_user
            p['password'] = new_pass
            _mark_u_dirty()
            logger.info(f"[DB] Credential sync: updated '{pname}' username/password.")

    # Fix: Flex sms was wrongly marked as reCAPTCHA panel — reCAPTCHA is commented
    # out on that panel; it uses a simple math captcha that the bot can solve.
    ps = _U.setdefault("panel_status", {})
    _flex = ps.get('Flex sms', {})
    if _flex.get('last_error') == 'Google reCAPTCHA — auto-login not possible':
        _flex['enabled']    = True
        _flex['last_error'] = ''
        ps['Flex sms']      = _flex
        _mark_u_dirty()
        logger.info("[DB] 'Flex sms' re-enabled — reCAPTCHA was commented out, math captcha is solvable.")

    # Migrate existing per-country services → global_services list
    if not _U["settings"].get("global_services"):
        seen: set = set()
        migrated: list = []
        for c in _N["countries"].values():
            for s in c.get("services", []):
                if s.lower() not in seen:
                    seen.add(s.lower())
                    migrated.append(s)
        if migrated:
            _U["settings"]["global_services"] = sorted(migrated)
            _mark_u_dirty()
            logger.info(f"[DB] Migrated {len(migrated)} services to global_services.")

    # Seed required_channels from config if DB list is still empty
    if not _U.get("required_channels"):
        try:
            from config import REQUIRED_CHANNELS as _cfg_chs
            if _cfg_chs:
                _U["required_channels"] = [
                    {"name": c["name"], "url": c["url"], "id": c["id"]}
                    for c in _cfg_chs
                ]
                _mark_u_dirty()
                logger.info(f"[DB] Seeded {len(_cfg_chs)} required_channels from config.")
        except Exception as _seed_err:
            logger.warning(f"[DB] Could not seed required_channels: {_seed_err}")

    orphans = _cleanup_orphan_countries()
    if orphans:
        logger.info(f"[DB] Cleanup: removed {orphans} orphan countries.")

    # Run cleanup once at startup to trim any old records left from previous runs
    try:
        m, d = _cleanup_old_records()
        if m or d:
            logger.info(f"[DB] Startup cleanup: removed {m} website_messages, {d} otp_deliveries (>7d)")
    except Exception as exc:
        logger.error(f"[DB] startup cleanup failed: {exc}")

    _atomic_write(_NUMBERS_FILE, _N)
    _atomic_write(_USERS_FILE, _U)

    threading.Thread(target=_flush_loop,   daemon=True, name="db-flush").start()
    threading.Thread(target=_cleanup_loop, daemon=True, name="db-cleanup").start()
    logger.info("[DB] Initialised — JSON file backend (numbers.json + users.json).")


# ══════════════════════════════════════════════════════════════════════════════
# Country / Number
# ══════════════════════════════════════════════════════════════════════════════

def _get_countries():
    with _lock:
        rows = sorted(_N["countries"].values(), key=lambda r: r["name"])
        return [(r["id"], r["name"]) for r in rows]


def _get_available_number_by_country(country_id):
    with _lock:
        cid = int(country_id)
        for rec in _N["numbers"].values():
            if int(rec["country_id"]) == cid and rec["used"] == 0:
                rec["used"] = 1
                _mark_n_dirty()
                return rec["number"]
        return None


def _get_available_numbers_by_country(country_id, limit=5):
    with _lock:
        cid    = int(country_id)
        result = []
        for rec in _N["numbers"].values():
            if int(rec["country_id"]) == cid and rec["used"] == 0:
                result.append(rec)
                if len(result) >= limit:
                    break
        for rec in result:
            rec["used"] = 1
        if result:
            _mark_n_dirty()
        return [r["number"] for r in result]


def _assign_number_to_user(user_id, number, country_id=None):
    with _lock:
        aid = _N["next_assignment_id"]
        _N["assignments"][str(aid)] = {
            "id": aid, "user_id": int(user_id),
            "number": number, "country_id": country_id,
            "assigned_at": _now(),
        }
        _N["next_assignment_id"] = aid + 1
        _mark_n_dirty()


def _get_user_by_number(number, max_age_seconds: int | None = None):
    with _lock:
        clean  = number.lstrip("+").lstrip("0")
        suffix = clean[-10:] if len(clean) >= 10 else clean
        now_ts = time.time()

        def _fresh(rec):
            if max_age_seconds is None:
                return True
            ts = _parse_ts(rec.get("assigned_at", ""))
            if ts <= 0:
                return False
            return (now_ts - ts) <= max_age_seconds

        best   = None
        best_id = -1
        for rec in _N["assignments"].values():
            rid = int(rec["id"])
            if rec["number"] == number and rid > best_id and _fresh(rec):
                best    = rec["user_id"]
                best_id = rid
        if best is not None:
            return best
        for rec in sorted(_N["assignments"].values(), key=lambda r: -int(r["id"])):
            if not _fresh(rec):
                continue
            stored = rec["number"].lstrip("+").lstrip("0")
            if (stored.endswith(suffix) or
                    suffix.endswith(stored[-10:] if len(stored) >= 10 else stored)):
                return rec["user_id"]
        return None


def _get_recent_user_by_number(number, max_age_seconds: int = 3600):
    """Return user_id only if number was assigned within max_age_seconds. Else None."""
    return _get_user_by_number(number, max_age_seconds=max_age_seconds)


def _get_numbers_count_by_country(country_id):
    with _lock:
        cid = int(country_id)
        total = avail = 0
        for rec in _N["numbers"].values():
            if int(rec["country_id"]) == cid:
                total += 1
                if rec["used"] == 0:
                    avail += 1
        return total, avail


def _get_all_country_counts():
    with _lock:
        counts: dict[int, list] = {}
        for rec in _N["numbers"].values():
            cid = int(rec["country_id"])
            if cid not in counts:
                counts[cid] = [0, 0]
            counts[cid][0] += 1
            if rec["used"] == 0:
                counts[cid][1] += 1
        return {cid: (t, a) for cid, (t, a) in counts.items()}


def _add_country(country_name):
    with _lock:
        for rec in _N["countries"].values():
            if rec["name"].lower() == country_name.lower():
                return False
        cid = _N["next_country_id"]
        _N["countries"][str(cid)] = {
            "id": cid, "name": country_name, "created_at": _now()
        }
        _N["next_country_id"] = cid + 1
        _mark_n_dirty()
        return True


def _add_numbers_to_country(country_id, numbers_list):
    with _lock:
        existing = {r["number"] for r in _N["numbers"].values()}
        added = 0
        for num in numbers_list:
            num = num.strip()
            if not num or num in existing:
                continue
            nid = _N["next_number_id"]
            _N["numbers"][str(nid)] = {
                "id": nid, "country_id": int(country_id),
                "number": num, "used": 0, "created_at": _now(),
            }
            _N["next_number_id"] = nid + 1
            existing.add(num)
            added += 1
        if added:
            _mark_n_dirty()
        return added


def _delete_number(number):
    with _lock:
        key = next((k for k, r in _N["numbers"].items() if r["number"] == number), None)
        if key:
            del _N["numbers"][key]
            _mark_n_dirty()
            return True
        return False


def _delete_all_numbers_from_country(country_id):
    with _lock:
        cid  = int(country_id)
        keys = [k for k, r in _N["numbers"].items() if int(r["country_id"]) == cid]
        for k in keys:
            del _N["numbers"][k]
        ckey = str(cid)
        if ckey in _N.get("countries", {}):
            del _N["countries"][ckey]
        if ckey in _N.get("country_otp_bonus", {}):
            del _N["country_otp_bonus"][ckey]
        a_keys = [k for k, r in _N.get("assignments", {}).items()
                  if r.get("country_id") is not None and int(r["country_id"]) == cid]
        for k in a_keys:
            del _N["assignments"][k]
        if keys or a_keys or ckey not in _N.get("countries", {}):
            _mark_n_dirty()
        return len(keys)


def _cleanup_orphan_countries():
    """Remove any country that has zero numbers (e.g., from older deletions)."""
    with _lock:
        used_cids = {int(r["country_id"]) for r in _N.get("numbers", {}).values()}
        orphans   = [k for k, c in list(_N.get("countries", {}).items())
                     if int(c["id"]) not in used_cids]
        for k in orphans:
            del _N["countries"][k]
            _N.get("country_otp_bonus", {}).pop(k, None)
        if orphans:
            _mark_n_dirty()
        return len(orphans)


def _delete_country(country_id):
    nd = _delete_all_numbers_from_country(country_id)
    with _lock:
        cid = int(country_id)
        key = str(cid)
        removed = False
        if key in _N["countries"]:
            del _N["countries"][key]
            removed = True
        if key in _N.get("country_otp_bonus", {}):
            del _N["country_otp_bonus"][key]
            removed = True
        a_keys = [k for k, r in _N.get("assignments", {}).items()
                  if r.get("country_id") is not None and int(r["country_id"]) == cid]
        for k in a_keys:
            del _N["assignments"][k]
        if removed or a_keys:
            _mark_n_dirty()
        return nd, (removed or bool(a_keys))


def _get_country_stats():
    with _lock:
        # O(numbers) precompute — avoids O(countries × numbers) nested scan
        counts: dict[int, list] = {}
        for r in _N["numbers"].values():
            cid = int(r["country_id"])
            if cid not in counts:
                counts[cid] = [0, 0]
            counts[cid][0] += 1
            if r["used"] == 0:
                counts[cid][1] += 1

        result = []
        for c in sorted(_N["countries"].values(), key=lambda x: x["name"]):
            cid = int(c["id"])
            pair = counts.get(cid, [0, 0])
            result.append((c["name"], pair[0], pair[1]))
        return result


def _get_country_id_by_name(country_name):
    with _lock:
        for rec in _N["countries"].values():
            if rec["name"].lower() == country_name.lower():
                return rec["id"]
        return None


def _reset_country_numbers(country_id):
    with _lock:
        cid = int(country_id)
        cnt = 0
        for rec in _N["numbers"].values():
            if int(rec["country_id"]) == cid and rec["used"] == 1:
                rec["used"] = 0
                cnt += 1
        # Drop assignments for this country so the 5-min window restarts fresh
        drop_keys = [k for k, rec in _N["assignments"].items()
                     if rec.get("country_id") is not None
                     and int(rec.get("country_id")) == cid]
        for k in drop_keys:
            del _N["assignments"][k]
        if cnt or drop_keys:
            _mark_n_dirty()
        return cnt


def _reset_all_numbers():
    with _lock:
        cnt = 0
        for rec in _N["numbers"].values():
            if rec["used"] == 1:
                rec["used"] = 0
                cnt += 1
        # Drop all assignments so the 5-min window restarts fresh
        had_assignments = bool(_N["assignments"])
        if had_assignments:
            _N["assignments"] = {}
        if cnt or had_assignments:
            _mark_n_dirty()
        return cnt


def _get_number_limit():
    with _lock:
        return int(_U["settings"].get("number_limit", "0"))


def _set_number_limit(limit: int):
    with _lock:
        _U["settings"]["number_limit"] = str(limit)
        _mark_u_dirty()


# ── Country OTP Bonus ──────────────────────────────────────────────────────────

def _get_country_otp_bonus(country_id: int):
    with _lock:
        val = _N["country_otp_bonus"].get(str(country_id))
        return float(val) if val is not None else None


def _set_country_otp_bonus(country_id: int, amount: float):
    with _lock:
        _N["country_otp_bonus"][str(country_id)] = amount
        _mark_n_dirty()


def _reset_country_otp_bonus(country_id: int):
    with _lock:
        _N["country_otp_bonus"].pop(str(country_id), None)
        _mark_n_dirty()


def _get_all_country_otp_bonuses() -> dict:
    with _lock:
        return {int(k): float(v) for k, v in _N["country_otp_bonus"].items()}


# ══════════════════════════════════════════════════════════════════════════════
# Country Services
# ══════════════════════════════════════════════════════════════════════════════

def _get_country_services(country_id: int) -> list:
    with _lock:
        country = _N["countries"].get(str(country_id))
        if not country:
            return []
        return list(country.get("services", []))


def _add_country_service(country_id: int, service_name: str) -> bool:
    with _lock:
        country = _N["countries"].get(str(country_id))
        if not country:
            return False
        services = country.setdefault("services", [])
        if any(s.lower() == service_name.lower() for s in services):
            return False
        services.append(service_name)
        _mark_n_dirty()
        return True


def _delete_country_service(country_id: int, service_idx: int) -> bool:
    with _lock:
        country = _N["countries"].get(str(country_id))
        if not country:
            return False
        services = country.get("services", [])
        if service_idx < 0 or service_idx >= len(services):
            return False
        services.pop(service_idx)
        _mark_n_dirty()
        return True


def _get_all_services() -> list:
    """Return sorted unique list of all services across all countries."""
    with _lock:
        seen = set()
        result = []
        for c in _N["countries"].values():
            for s in c.get("services", []):
                if s not in seen:
                    seen.add(s)
                    result.append(s)
        return sorted(result)


def _get_countries_by_service(service_name: str) -> list:
    """Return list of (id, name) for countries that have this service."""
    with _lock:
        result = []
        for c in sorted(_N["countries"].values(), key=lambda x: x["name"]):
            if any(s.lower() == service_name.lower() for s in c.get("services", [])):
                result.append((c["id"], c["name"]))
        return result


def _get_global_services() -> list:
    """Return the admin-defined global services list."""
    with _lock:
        return list(_U["settings"].get("global_services", []))


def _add_global_service(name: str) -> bool:
    """Add a service to the global list. Returns True if added, False if duplicate."""
    with _lock:
        services = _U["settings"].setdefault("global_services", [])
        if any(s.lower() == name.lower() for s in services):
            return False
        services.append(name)
        _mark_u_dirty()
        return True


def _remove_global_service(name: str) -> bool:
    """Remove a service from the global list and from all country mappings."""
    with _lock:
        services = _U["settings"].get("global_services", [])
        match = next((s for s in services if s.lower() == name.lower()), None)
        if match is None:
            return False
        services.remove(match)
        for c in _N["countries"].values():
            c_svcs = c.get("services", [])
            to_del = [s for s in c_svcs if s.lower() == name.lower()]
            for s in to_del:
                c_svcs.remove(s)
        _mark_u_dirty()
        _mark_n_dirty()
        return True


def _unmap_service_from_country(service_name: str, country_id: int) -> bool:
    """Remove a service mapping from a specific country."""
    with _lock:
        country = _N["countries"].get(str(country_id))
        if not country:
            return False
        services = country.get("services", [])
        match = next((s for s in services if s.lower() == service_name.lower()), None)
        if match is None:
            return False
        services.remove(match)
        _mark_n_dirty()
        return True


# ══════════════════════════════════════════════════════════════════════════════
# Admin
# ══════════════════════════════════════════════════════════════════════════════

def _add_admin(username):
    with _lock:
        if username in _U["admins"]:
            return False
        _U["admins"][username] = {"username": username, "added_at": _now()}
        _mark_u_dirty()
        _invalidate_admin_cache()
        return True


def _add_admin_by_uid(uid: int):
    with _lock:
        row = _U["users"].get(str(uid))
        if not row:
            return False, None, "user_not_found"
        uname = row.get("username") or f"id_{uid}"
        if uname in _U["admins"]:
            return False, dict(row), "already_admin"
        _U["admins"][uname] = {"username": uname, "added_at": _now()}
        _mark_u_dirty()
        _invalidate_admin_cache()
        return True, dict(row), "ok"


def _remove_admin(username):
    from config import PROTECTED_ADMINS as PA, PROTECTED_ADMIN_IDS as PAI
    if username in PA:
        return False, "Cannot remove protected admin"
    with _lock:
        user_row = next(
            (r for r in _U["users"].values() if r.get("username") == username), None
        )
        if user_row and int(user_row["user_id"]) in PAI:
            return False, "Cannot remove protected admin"
        if username in _U["admins"]:
            del _U["admins"][username]
            _mark_u_dirty()
            _invalidate_admin_cache()
            return True, "Admin removed successfully"
        return False, "Admin not found"


def _get_all_admins():
    with _lock:
        return list(_U["admins"].keys())


def _get_all_admins_with_details():
    with _lock:
        result = []
        for uname, arec in sorted(
            _U["admins"].items(), key=lambda x: x[1]["added_at"], reverse=True
        ):
            user_row = next(
                (r for r in _U["users"].values() if r.get("username") == uname), {}
            )
            result.append({
                "username":   uname,
                "added_at":   arec["added_at"],
                "user_id":    user_row.get("user_id"),
                "first_name": user_row.get("first_name"),
                "last_name":  user_row.get("last_name"),
            })
        return result


def _is_admin(username, user_id=None):
    from config import PROTECTED_ADMIN_IDS as PAI
    if user_id and int(user_id) in PAI:
        return True
    if not username:
        return False
    return username in _get_cached_admins()


# ══════════════════════════════════════════════════════════════════════════════
# User
# ══════════════════════════════════════════════════════════════════════════════

def _add_user(user_id, username, first_name, last_name, referred_by=None):
    with _lock:
        uid      = str(user_id)
        ref_code = _generate_referral_code(user_id)
        existing = _U["users"].get(uid)
        if existing:
            existing.update({
                "username":     username,
                "first_name":   first_name,
                "last_name":    last_name,
                "referral_code": ref_code,
            })
        else:
            _U["users"][uid] = {
                "user_id":       int(user_id),
                "username":      username,
                "first_name":    first_name,
                "last_name":     last_name,
                "balance":       0.0,
                "referral_code": ref_code,
                "referred_by":   referred_by,
                "is_verified":   0,
                "verified_at":   None,
                "joined_at":     _now(),
            }
        _mark_u_dirty()


def _get_all_users():
    with _lock:
        return [int(r["user_id"]) for r in _U["users"].values()]


def _get_all_users_with_info():
    with _lock:
        return [
            (r["user_id"], r.get("username"), r.get("first_name"),
             r.get("last_name"), r.get("joined_at"))
            for r in _U["users"].values()
        ]


def _get_user_count():
    with _lock:
        return len(_U["users"])


def _get_user_stats_summary() -> dict:
    with _lock:
        total_users         = len(_U["users"])
        total_otp_count     = len(_U["otp_bonus_log"])
        total_otp_amount    = sum(float(r.get("amount", 0))
                                  for r in _U["otp_bonus_log"].values())
        total_referrals     = len(_U["referral_log"])
        total_ref_amount    = sum(float(r.get("bonus_amount", 0))
                                  for r in _U["referral_log"].values())
        total_withdrawals   = len(_U["withdraw_requests"])
        total_withdraw_amt  = sum(float(r.get("amount", 0))
                                  for r in _U["withdraw_requests"].values())
        user_balances = [
            (r.get("username") or r.get("first_name") or str(r["user_id"]),
             float(r.get("balance", 0)))
            for r in _U["users"].values()
        ]
        top_earners = sorted(user_balances, key=lambda x: x[1], reverse=True)[:5]
        return {
            "total_users":         total_users,
            "total_otp_count":     total_otp_count,
            "total_otp_amount":    round(total_otp_amount, 2),
            "total_referrals":     total_referrals,
            "total_ref_amount":    round(total_ref_amount, 2),
            "total_withdrawals":   total_withdrawals,
            "total_withdraw_amt":  round(total_withdraw_amt, 2),
            "top_earners":         top_earners,
        }


def _is_user_verified(user_id):
    with _lock:
        row = _U["users"].get(str(user_id))
        return bool(row and row.get("is_verified"))


# ══════════════════════════════════════════════════════════════════════════════
# Owner Panel — Admin-button enable/disable persistence
# ══════════════════════════════════════════════════════════════════════════════

def _get_disabled_admin_buttons():
    """Return the list of admin-button labels currently disabled by the owner."""
    with _lock:
        raw = _U.get("disabled_admin_buttons", [])
        if not isinstance(raw, list):
            return []
        return [str(x) for x in raw]


def _set_admin_button_disabled(name: str, disabled: bool) -> bool:
    """Mark an admin-panel button as disabled (True) or enabled (False).
    Returns True if the state actually changed."""
    name = str(name).strip()
    if not name:
        return False
    with _lock:
        cur = list(_U.get("disabled_admin_buttons", []))
        if not isinstance(cur, list):
            cur = []
        cur_set = set(cur)
        was_disabled = name in cur_set
        if disabled and not was_disabled:
            cur_set.add(name)
            _U["disabled_admin_buttons"] = sorted(cur_set)
            _mark_u_dirty()
            return True
        if (not disabled) and was_disabled:
            cur_set.discard(name)
            _U["disabled_admin_buttons"] = sorted(cur_set)
            _mark_u_dirty()
            return True
        return False


# ══════════════════════════════════════════════════════════════════════════════
# Panels
# ══════════════════════════════════════════════════════════════════════════════

def _get_panels():
    with _lock:
        return list(_U["panels"].values())


def _get_panel_by_name(name):
    with _lock:
        return _U["panels"].get(name)


def _update_panel_credentials(name, username, password):
    with _lock:
        if name in _U["panels"]:
            _U["panels"][name]["username"] = username
            _U["panels"][name]["password"] = password
            _mark_u_dirty()
            return True
        return False


def _get_all_panel_statuses():
    with _lock:
        return list(_U["panel_status"].values())


def _set_panel_enabled(panel_name, enabled):
    """Save a panel's enabled/disabled state to JSON (persistent)."""
    with _lock:
        rec = _U["panel_status"].setdefault(panel_name, {
            "panel_name":       panel_name,
            "logged_in":        0,
            "last_login_at":    None,
            "last_fetch_at":    None,
            "last_fetch_count": None,
            "last_error":       None,
        })
        rec["enabled"] = bool(enabled)
        _mark_u_dirty()


def _is_panel_enabled(panel_name) -> bool:
    """Check whether a panel is currently enabled. Defaults to True if unset."""
    with _lock:
        rec = _U.get("panel_status", {}).get(panel_name)
        if not rec:
            return True
        return bool(rec.get("enabled", True))


def _get_panel_interval(panel_name: str) -> int | None:
    """Return the stored polling interval (seconds) for a panel, or None if not set."""
    with _lock:
        p = _U["panels"].get(panel_name)
        if p and "interval" in p:
            return int(p["interval"])
        return None


def _set_panel_interval(panel_name: str, seconds: int):
    """Persist the polling interval (seconds) for a panel."""
    with _lock:
        p = _U["panels"].get(panel_name)
        if p is not None:
            p["interval"] = int(seconds)
            _mark_u_dirty()


def _get_panel_retry_interval(panel_name: str) -> int | None:
    """Return the stored login-retry interval (seconds) for a panel, or None if not set."""
    with _lock:
        p = _U["panels"].get(panel_name)
        if p and "retry_interval" in p:
            return int(p["retry_interval"])
        return None


def _set_panel_retry_interval(panel_name: str, seconds: int):
    """Persist the login-retry interval (seconds) for a panel."""
    with _lock:
        p = _U["panels"].get(panel_name)
        if p is not None:
            p["retry_interval"] = int(seconds)
            _mark_u_dirty()


# ══════════════════════════════════════════════════════════════════════════════
# Extra Groups
# ══════════════════════════════════════════════════════════════════════════════

def _add_extra_group(chat_id: str, title: str):
    with _lock:
        groups = _U.setdefault("extra_groups", [])
        for g in groups:
            if g["chat_id"] == str(chat_id):
                return
        next_id = max((g["id"] for g in groups), default=0) + 1
        groups.append({
            "id":       next_id,
            "chat_id":  str(chat_id),
            "title":    title,
            "added_at": _now(),
        })
        _mark_u_dirty()


def _remove_extra_group(chat_id: str):
    with _lock:
        groups = _U.setdefault("extra_groups", [])
        _U["extra_groups"] = [g for g in groups if g["chat_id"] != str(chat_id)]
        _mark_u_dirty()


def _get_all_extra_groups() -> list:
    with _lock:
        return list(_U.get("extra_groups", []))


# ── Group member tracking (records every member the bot sees in a group) ─────
# Telegram Bot API does NOT expose a way to list ALL members of a group.
# What we CAN do: record every user the bot observes via chat_member updates,
# message updates, or get_chat_administrators. Stored as:
#   _U["group_members"][chat_id_str][user_id_str] = {...}
def _record_group_member(chat_id, user, status: str = "member"):
    """Record/refresh a single user observed in a group."""
    if user is None:
        return
    with _lock:
        gm        = _U.setdefault("group_members", {})
        bucket    = gm.setdefault(str(chat_id), {})
        uid       = str(getattr(user, "id", user))
        existing  = bucket.get(uid, {})
        bucket[uid] = {
            "user_id":     getattr(user, "id", None),
            "username":    getattr(user, "username", None) or "",
            "first_name":  getattr(user, "first_name", None) or "",
            "last_name":   getattr(user, "last_name", None) or "",
            "is_bot":      bool(getattr(user, "is_bot", False)),
            "language":    getattr(user, "language_code", None) or "",
            "status":      status or existing.get("status") or "member",
            "first_seen":  existing.get("first_seen") or _now(),
            "last_seen":   _now(),
        }
        _mark_u_dirty()


def _remove_group_member(chat_id, user_id):
    with _lock:
        gm     = _U.setdefault("group_members", {})
        bucket = gm.get(str(chat_id))
        if bucket and str(user_id) in bucket:
            del bucket[str(user_id)]
            _mark_u_dirty()


def _get_group_members(chat_id) -> list:
    """Return all tracked members of a group, sorted by first_seen."""
    with _lock:
        gm     = _U.get("group_members", {})
        bucket = gm.get(str(chat_id), {})
        members = list(bucket.values())
        members.sort(key=lambda m: (m.get("first_seen") or "", m.get("user_id") or 0))
        return members


def _get_group_member_count(chat_id) -> int:
    with _lock:
        gm     = _U.get("group_members", {})
        return len(gm.get(str(chat_id), {}))


def _update_panel_status(panel_name, logged_in, msg_count=None, error=None):
    with _lock:
        rec = _U["panel_status"].setdefault(panel_name, {
            "panel_name":       panel_name,
            "logged_in":        0,
            "last_login_at":    None,
            "last_fetch_at":    None,
            "last_fetch_count": None,
            "last_error":       None,
        })
        rec["logged_in"]  = 1 if logged_in else 0
        rec["last_error"] = error
        if logged_in:
            rec["last_login_at"] = _now()
        if msg_count is not None:
            rec["last_fetch_count"] = msg_count
            rec["last_fetch_at"]    = _now()
        _mark_u_dirty()


# ══════════════════════════════════════════════════════════════════════════════
# Settings
# ══════════════════════════════════════════════════════════════════════════════

def _get_setting(key, default=""):
    with _lock:
        return _U["settings"].get(key, default)


def _set_setting(key, value):
    with _lock:
        _U["settings"][key] = str(value)
        _mark_u_dirty()


def _get_referral_settings():
    enabled = _get_setting("referral_enabled", "1") == "1"
    bonus   = float(_get_setting("referral_bonus", "10.0"))
    return {"enabled": enabled, "bonus": bonus}


def _set_referral_bonus(amount):
    _set_setting("referral_bonus", amount)


def _toggle_referral(enabled: bool):
    _set_setting("referral_enabled", "1" if enabled else "0")


def _get_min_withdraw():
    return float(_get_setting("min_withdraw", "50.0"))


def _set_min_withdraw(amount):
    _set_setting("min_withdraw", amount)


def _get_otp_bonus_settings():
    enabled     = _get_setting("otp_bonus_enabled", "1") == "1"
    amount      = float(_get_setting("otp_bonus_amount", "2.0"))
    daily_limit = int(_get_setting("otp_bonus_daily_limit", "10"))
    return {"enabled": enabled, "amount": amount, "daily_limit": daily_limit}


def _toggle_otp_bonus(enabled: bool):
    _set_setting("otp_bonus_enabled", "1" if enabled else "0")


def _set_otp_bonus_amount(amount: float):
    _set_setting("otp_bonus_amount", str(amount))


def _set_otp_daily_limit(limit: int):
    _set_setting("otp_bonus_daily_limit", str(limit))


# ══════════════════════════════════════════════════════════════════════════════
# Balance & Referral
# ══════════════════════════════════════════════════════════════════════════════

def _get_user_balance(user_id):
    with _lock:
        row = _U["users"].get(str(user_id))
        return round(float(row["balance"]), 4) if row else 0.0


def _update_user_balance(user_id, amount):
    with _lock:
        row = _U["users"].get(str(user_id))
        if row:
            row["balance"] = round(float(row["balance"]) + float(amount), 4)
            _mark_u_dirty()


def _set_user_balance(user_id, amount):
    with _lock:
        row = _U["users"].get(str(user_id))
        if row:
            row["balance"] = round(float(amount), 4)
            _mark_u_dirty()


def _get_user_by_ref_code(ref_code):
    with _lock:
        for row in _U["users"].values():
            if row.get("referral_code") == ref_code:
                return row["user_id"]
        return None


def _get_user_referral_code(user_id):
    with _lock:
        row = _U["users"].get(str(user_id))
        if row:
            if row.get("referral_code"):
                return row["referral_code"]
            code = _generate_referral_code(user_id)
            row["referral_code"] = code
            _mark_u_dirty()
            return code
        return None


def _get_user_info_by_id(user_id):
    with _lock:
        row = _U["users"].get(str(user_id))
        if row:
            return {
                "user_id":       row["user_id"],
                "username":      row.get("username"),
                "first_name":    row.get("first_name"),
                "balance":       float(row.get("balance", 0)),
                "referral_code": row.get("referral_code"),
            }
        return None


def _credit_referral(referrer_id, referred_id, bonus_amount):
    with _lock:
        for rec in _U["referral_log"].values():
            if int(rec["referred_id"]) == int(referred_id):
                return False
        rid = _U["next_ref_id"]
        _U["referral_log"][str(rid)] = {
            "id":           rid,
            "referrer_id":  int(referrer_id),
            "referred_id":  int(referred_id),
            "bonus_amount": float(bonus_amount),
            "created_at":   _now(),
        }
        _U["next_ref_id"] = rid + 1
        _mark_u_dirty()
    _update_user_balance(referrer_id, bonus_amount)
    return True


def _get_referral_count(user_id):
    with _lock:
        uid = int(user_id)
        return sum(1 for r in _U["referral_log"].values()
                   if int(r["referrer_id"]) == uid)


def _get_referral_total_earned(user_id):
    with _lock:
        uid = int(user_id)
        return round(
            sum(float(r["bonus_amount"]) for r in _U["referral_log"].values()
                if int(r["referrer_id"]) == uid), 4
        )


def _get_top_referrers(limit=10):
    with _lock:
        counts: dict[int, dict] = {}
        for rec in _U["referral_log"].values():
            rid = int(rec["referrer_id"])
            if rid not in counts:
                counts[rid] = {"count": 0, "earned": 0.0}
            counts[rid]["count"]  += 1
            counts[rid]["earned"] += float(rec["bonus_amount"])
        result = []
        for uid, stats in sorted(counts.items(), key=lambda x: -x[1]["count"])[:limit]:
            user_row = _U["users"].get(str(uid), {})
            result.append({
                "user_id":    uid,
                "username":   user_row.get("username"),
                "first_name": user_row.get("first_name"),
                "count":      stats["count"],
                "earned":     round(stats["earned"], 4),
            })
        return result


def _get_top_users_detailed(limit: int = 5) -> list[dict]:
    """Return top `limit` users sorted by numbers used (assignments).
    Each entry contains user_id, display_name, numbers_used, msgs_received,
    referral_count."""
    with _lock:
        # Count assignments per user
        numbers_used: dict[int, int] = {}
        for rec in _N.get("assignments", {}).values():
            uid = int(rec.get("user_id", 0))
            if uid:
                numbers_used[uid] = numbers_used.get(uid, 0) + 1

        # Count OTP bonus entries per user (msgs received)
        msgs_received: dict[int, int] = {}
        for rec in _U.get("otp_bonus_log", {}).values():
            uid = int(rec.get("user_id", 0))
            if uid:
                msgs_received[uid] = msgs_received.get(uid, 0) + 1

        # Count referrals per user
        referral_count: dict[int, int] = {}
        for rec in _U.get("referral_log", {}).values():
            uid = int(rec.get("referrer_id", 0))
            if uid:
                referral_count[uid] = referral_count.get(uid, 0) + 1

        # Build list from all known users, sorted by numbers_used descending
        all_uids = set(_U.get("users", {}).keys())
        rows = []
        for uid_str in all_uids:
            uid = int(uid_str)
            user_row = _U["users"].get(uid_str, {})
            display = (
                user_row.get("username")
                or user_row.get("first_name")
                or f"ID:{uid}"
            )
            rows.append({
                "user_id":        uid,
                "display_name":   display,
                "numbers_used":   numbers_used.get(uid, 0),
                "msgs_received":  msgs_received.get(uid, 0),
                "referral_count": referral_count.get(uid, 0),
                "balance":        float(user_row.get("balance", 0.0)),
            })

        rows.sort(key=lambda r: (r["msgs_received"], r["balance"]), reverse=True)
        return rows[:limit]


# ══════════════════════════════════════════════════════════════════════════════
# Withdraw
# ══════════════════════════════════════════════════════════════════════════════

def _create_withdraw_request(user_id, amount, method, account):
    with _lock:
        wid = _U["next_withdraw_id"]
        _U["withdraw_requests"][str(wid)] = {
            "id": wid, "user_id": int(user_id),
            "amount": float(amount), "method": method,
            "account": account, "status": "pending",
            "created_at": _now(),
        }
        _U["next_withdraw_id"] = wid + 1
        _mark_u_dirty()
        return wid


def _get_pending_withdraws():
    with _lock:
        pending = [r for r in _U["withdraw_requests"].values()
                   if r["status"] == "pending"]
        pending.sort(key=lambda r: r["created_at"])
        result = []
        for r in pending:
            user_row = _U["users"].get(str(r["user_id"]), {})
            result.append({
                "id":         r["id"],
                "user_id":    r["user_id"],
                "username":   user_row.get("username"),
                "first_name": user_row.get("first_name"),
                "amount":     float(r["amount"]),
                "method":     r["method"],
                "account":    r["account"],
                "created_at": r["created_at"],
            })
        return result


def _get_withdraw_request_by_id(request_id):
    with _lock:
        r = _U["withdraw_requests"].get(str(request_id))
        if r:
            return {
                "id":      r["id"],
                "user_id": r["user_id"],
                "amount":  float(r["amount"]),
                "method":  r["method"],
                "account": r["account"],
                "status":  r["status"],
            }
        return None


def _update_withdraw_status(request_id, status):
    with _lock:
        rec = _U["withdraw_requests"].get(str(request_id))
        if rec:
            rec["status"] = status
            _mark_u_dirty()


# ══════════════════════════════════════════════════════════════════════════════
# OTP Bonus
# ══════════════════════════════════════════════════════════════════════════════

def _has_otp_bonus_received(delivery_key: str) -> bool:
    with _lock:
        return any(r["delivery_key"] == delivery_key
                   for r in _U["otp_bonus_log"].values())


def _get_otp_bonus_today_count(user_id: int) -> int:
    with _lock:
        uid   = int(user_id)
        today = _today()
        return sum(1 for r in _U["otp_bonus_log"].values()
                   if int(r["user_id"]) == uid and r["created_at"].startswith(today))


def _get_user_otp_bonus_stats(user_id: int) -> dict:
    with _lock:
        uid   = int(user_id)
        today = _today()
        tc = te = ac = ae = 0
        for r in _U["otp_bonus_log"].values():
            if int(r["user_id"]) != uid:
                continue
            a = float(r["amount"])
            ac += 1; ae += a
            if r["created_at"].startswith(today):
                tc += 1; te += a
        return {
            "today_count":  tc, "today_earned": round(te, 4),
            "total_count":  ac, "total_earned": round(ae, 4),
        }


def _record_otp_bonus(user_id: int, delivery_key: str, amount: float) -> bool:
    with _lock:
        if any(r["delivery_key"] == delivery_key
               for r in _U["otp_bonus_log"].values()):
            return False
        bid = _U["next_otp_bonus_id"]
        _U["otp_bonus_log"][str(bid)] = {
            "id": bid, "user_id": int(user_id),
            "delivery_key": delivery_key,
            "amount": float(amount), "created_at": _now(),
        }
        _U["next_otp_bonus_id"] = bid + 1
        _mark_u_dirty()
    _update_user_balance(user_id, amount)
    return True


def _get_effective_otp_bonus(number: str, global_amount: float) -> float:
    with _lock:
        assignment = None
        for rec in sorted(_N["assignments"].values(), key=lambda r: -int(r["id"])):
            if rec["number"] == number or rec["number"] == "+" + number:
                assignment = rec
                break
        if not assignment or not assignment.get("country_id"):
            return global_amount
        val = _N["country_otp_bonus"].get(str(assignment["country_id"]))
        return float(val) if val is not None else global_amount


# ══════════════════════════════════════════════════════════════════════════════
# OTP Deliveries
# ══════════════════════════════════════════════════════════════════════════════

def _is_otp_delivered(delivery_key: str) -> bool:
    with _lock:
        return delivery_key in _U["otp_deliveries"]


def _mark_otp_delivered(delivery_key: str):
    with _lock:
        if delivery_key not in _U["otp_deliveries"]:
            _U["otp_deliveries"][delivery_key] = _now()
            _mark_u_dirty()


# ══════════════════════════════════════════════════════════════════════════════
# Website Messages  (SMS log — capped at 10 000 to keep RAM reasonable)
# ══════════════════════════════════════════════════════════════════════════════

_WEBSITE_PATTERNS = [
    (re.compile(r"\b(facebook|fb)\b", re.I),              "Facebook"),
    (re.compile(r"\b(instagram|ig)\b", re.I),             "Instagram"),
    (re.compile(r"\b(whatsapp)\b", re.I),                 "WhatsApp"),
    (re.compile(r"\b(telegram)\b", re.I),                 "Telegram"),
    (re.compile(r"\b(google|gmail|youtube)\b", re.I),     "Google"),
    (re.compile(r"\b(twitter|x\.com)\b", re.I),           "Twitter/X"),
    (re.compile(r"\b(tiktok)\b", re.I),                   "TikTok"),
    (re.compile(r"\b(snapchat)\b", re.I),                 "Snapchat"),
    (re.compile(r"\b(amazon)\b", re.I),                   "Amazon"),
    (re.compile(r"\b(netflix)\b", re.I),                  "Netflix"),
    (re.compile(r"\b(microsoft|outlook|hotmail)\b", re.I),"Microsoft"),
    (re.compile(r"\b(apple|icloud|itunes)\b", re.I),      "Apple"),
    (re.compile(r"\b(paypal)\b", re.I),                   "PayPal"),
    (re.compile(r"\b(uber)\b", re.I),                     "Uber"),
    (re.compile(r"\b(linkedin)\b", re.I),                 "LinkedIn"),
    (re.compile(r"\b(binance)\b", re.I),                  "Binance"),
    (re.compile(r"\b(coinbase)\b", re.I),                 "Coinbase"),
    (re.compile(r"\b(discord)\b", re.I),                  "Discord"),
    (re.compile(r"\b(spotify)\b", re.I),                  "Spotify"),
    (re.compile(r"\b(shopify)\b", re.I),                  "Shopify"),
    (re.compile(r"\b(alibaba|aliexpress)\b", re.I),       "Alibaba"),
    (re.compile(r"\b(lazada)\b", re.I),                   "Lazada"),
    (re.compile(r"\b(grab)\b", re.I),                     "Grab"),
    (re.compile(r"\b(airbnb)\b", re.I),                   "Airbnb"),
    (re.compile(r"\b(ebay)\b", re.I),                     "eBay"),
]


def _detect_website(message: str) -> str:
    for pat, name in _WEBSITE_PATTERNS:
        if pat.search(message):
            return name
    m = re.search(r"(?:https?://)?(?:www\.)?([a-zA-Z0-9\-]+)\.[a-z]{2,}", message)
    if m:
        d = m.group(1).capitalize()
        if len(d) > 2:
            return d
    return "Unknown"


_MAX_WEBSITE_MSGS    = 10_000
_latest_msg_cache    = {"record": None}
_panel_latest_cache  = {}


def _save_website_message(
    website_name: str, number: str, message: str,
    msg_timestamp, country, otp, panel_name: str = "",
):
    with _lock:
        mid = _U["next_website_msg_id"]
        rec = {
            "id":            mid,
            "website_name":  website_name,
            "number":        number,
            "message":       message,
            "msg_timestamp": msg_timestamp,
            "country":       country or "",
            "otp":           otp or "",
            "received_at":   _now(),
            "panel_name":    panel_name,
        }
        _U["website_messages"][str(mid)] = rec
        _U["next_website_msg_id"] = mid + 1

        _latest_msg_cache["record"] = rec
        if panel_name:
            _panel_latest_cache[panel_name] = rec

        # Prune oldest entries when over cap
        if len(_U["website_messages"]) > _MAX_WEBSITE_MSGS:
            oldest_key = str(mid - _MAX_WEBSITE_MSGS)
            _U["website_messages"].pop(oldest_key, None)

        _mark_u_dirty()


def _get_latest_website_message():
    with _lock:
        cached = _latest_msg_cache.get("record")
        if cached:
            return cached
        if not _U["website_messages"]:
            return None
        latest = max(_U["website_messages"].values(), key=lambda r: int(r["id"]))
        _latest_msg_cache["record"] = latest
        return latest


def _get_latest_panel_message(panel_name: str):
    with _lock:
        cached = _panel_latest_cache.get(panel_name)
        if cached:
            return cached
        recs = [r for r in _U["website_messages"].values()
                if r.get("panel_name") == panel_name]
        if not recs:
            return None
        latest = max(recs, key=lambda r: int(r["id"]))
        _panel_latest_cache[panel_name] = latest
        return latest


# ══════════════════════════════════════════════════════════════════════════════
# Reset All User Data
# ══════════════════════════════════════════════════════════════════════════════

def _reset_all_user_data(exclude_user_id: int = None) -> dict:
    with _lock:
        today   = _today()
        summary = {
            "users":            len(_U["users"]),
            "referral_count":   len(_U["referral_log"]),
            "referral_income":  sum(float(r["bonus_amount"])
                                    for r in _U["referral_log"].values()),
            "otp_income_today": sum(float(r["amount"])
                                    for r in _U["otp_bonus_log"].values()
                                    if r["created_at"].startswith(today)),
            "otp_income_total": sum(float(r["amount"])
                                    for r in _U["otp_bonus_log"].values()),
            "withdraw_requests":  len(_U["withdraw_requests"]),
            "number_assignments": len(_N["assignments"]),
        }
        for row in _U["users"].values():
            row["balance"]    = 0.0
            row["referred_by"] = None
        _U["referral_log"]     = {}
        _U["withdraw_requests"] = {}
        _U["otp_bonus_log"]    = {}
        _U["otp_deliveries"]   = {}
        _U["website_messages"] = {}
        _U["next_ref_id"]          = 1
        _U["next_withdraw_id"]     = 1
        _U["next_otp_bonus_id"]    = 1
        _U["next_delivery_id"]     = 1
        _U["next_website_msg_id"]  = 1
        _N["assignments"]          = {}
        _N["next_assignment_id"]   = 1
        _latest_msg_cache["record"] = None
        _panel_latest_cache.clear()
        _mark_u_dirty()
        _mark_n_dirty()
    return summary


# ══════════════════════════════════════════════════════════════════════════════
# Excel / ZIP export
# ══════════════════════════════════════════════════════════════════════════════

def generate_users_excel():
    users     = _get_all_users_with_info()
    wb        = openpyxl.Workbook()
    ws        = wb.active
    ws.title  = "Users"
    hdr_fill  = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", size=12)
    ctr_align = Alignment(horizontal="center", vertical="center")
    headers    = ["#", "User ID", "Username", "First Name", "Last Name", "Joined At"]
    col_widths = [5, 15, 20, 20, 20, 25]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    for ri, (uid, uname, fname, lname, joined) in enumerate(users, 2):
        fill = PatternFill(
            start_color="DCE6F1" if ri % 2 == 0 else "FFFFFF",
            end_color  ="DCE6F1" if ri % 2 == 0 else "FFFFFF",
            fill_type  ="solid",
        )
        for ci, val in enumerate(
            [ri-1, uid, f"@{uname}" if uname else "—",
             fname or "—", lname or "—", joined or "—"], 1
        ):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.alignment = ctr_align
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_user_stats_excel():
    """Generate an Excel file with user statistics summary and top earners."""
    s = _get_user_stats_summary()
    wb        = openpyxl.Workbook()
    hdr_fill  = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", size=12)
    ctr_align = Alignment(horizontal="center", vertical="center")
    alt_fill  = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    summary_rows = [
        ("Metric", "Value"),
        ("Total Users",              s["total_users"]),
        ("Total OTP Bonuses",        s["total_otp_count"]),
        ("Total OTP Amount (BDT)",   s["total_otp_amount"]),
        ("Total Referrals",          s["total_referrals"]),
        ("Total Referral Amount (BDT)", s["total_ref_amount"]),
        ("Total Withdraw Requests",  s["total_withdrawals"]),
        ("Total Withdraw Amount (BDT)", s["total_withdraw_amt"]),
    ]
    col_widths = [35, 20]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=ci).column_letter].width = w
    for ri, (label, val) in enumerate(summary_rows, 1):
        c1 = ws1.cell(row=ri, column=1, value=label)
        c2 = ws1.cell(row=ri, column=2, value=val)
        for c in (c1, c2):
            c.alignment = ctr_align
            if ri == 1:
                c.font = hdr_font
                c.fill = hdr_fill
            elif ri % 2 == 0:
                c.fill = alt_fill

    # ── Sheet 2: Top Earners ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Top Earners")
    earner_headers = ["Rank", "Username / Name", "Balance (BDT)"]
    earner_widths  = [8, 30, 20]
    for ci, (h, w) in enumerate(zip(earner_headers, earner_widths), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr_align
        ws2.column_dimensions[cell.column_letter].width = w
    for ri, (name, bal) in enumerate(s["top_earners"], 2):
        display = f"@{name}" if name and not str(name).isdigit() else str(name)
        fill = alt_fill if ri % 2 == 0 else PatternFill(fill_type=None)
        for ci, val in enumerate([ri - 1, display, bal], 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.alignment = ctr_align
            if ri % 2 == 0:
                c.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_all_data_as_zip() -> tuple:
    stamp     = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    hdr_fill  = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    ctr_align = Alignment(horizontal="center", vertical="center")

    def _make_sheet(wb, title, headers, rows, col_widths):
        ws = wb.create_sheet(title=title)
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr_align
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 18
        for ri, row in enumerate(rows, 2):
            alt = PatternFill(
                start_color="DCE6F1" if ri % 2 == 0 else "FFFFFF",
                end_color  ="DCE6F1" if ri % 2 == 0 else "FFFFFF",
                fill_type  ="solid",
            )
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = alt; cell.alignment = ctr_align

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    with _lock:
        user_rows = [
            [r["user_id"], r.get("username"), r.get("first_name"), r.get("last_name"),
             r.get("balance", 0.0), r.get("referral_code"), r.get("referred_by"),
             r.get("is_verified", 0), r.get("joined_at")]
            for r in _U["users"].values()
        ]
        ref_rows = [
            [r["id"], r["referrer_id"], r["referred_id"],
             r["bonus_amount"], r["created_at"]]
            for r in _U["referral_log"].values()
        ]
        wr_rows = [
            [r["id"], r["user_id"], r["amount"], r["method"],
             r["account"], r["status"], r["created_at"]]
            for r in _U["withdraw_requests"].values()
        ]
        otp_rows = [
            [r["id"], r["user_id"], r["delivery_key"],
             r["amount"], r["created_at"]]
            for r in _U["otp_bonus_log"].values()
        ]
        assign_rows = [
            [r["id"], r["user_id"], r["number"],
             r.get("country_id"), r["assigned_at"]]
            for r in sorted(_N["assignments"].values(), key=lambda x: int(x["id"]))
        ]
        sms_rows = [
            [r["id"], r["website_name"], r["number"], r["message"],
             r.get("otp"), r.get("country"), r["received_at"]]
            for r in sorted(_U["website_messages"].values(),
                            key=lambda x: -int(x["id"]))[:5000]
        ]

    _make_sheet(wb, "Users",
        ["User ID","Username","First Name","Last Name","Balance (৳)",
         "Referral Code","Referred By","Verified","Joined At"],
        user_rows, [14,18,16,16,12,14,14,10,22])
    _make_sheet(wb, "Referral Log",
        ["#","Referrer ID","Referred ID","Bonus (৳)","Date"],
        ref_rows, [6,14,14,12,22])
    _make_sheet(wb, "Withdraw Requests",
        ["#","User ID","Amount (৳)","Method","Account","Status","Date"],
        wr_rows, [6,14,12,10,20,10,22])
    _make_sheet(wb, "OTP Bonus Log",
        ["#","User ID","Delivery Key","Amount (৳)","Date"],
        otp_rows, [6,14,30,12,22])
    _make_sheet(wb, "Number Assignments",
        ["#","User ID","Number","Country ID","Assigned At"],
        assign_rows, [6,14,18,12,22])
    _make_sheet(wb, "SMS Log",
        ["#","Website","Number","Message","OTP","Country","Received At"],
        sms_rows, [6,18,18,50,12,12,22])

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"backup_{stamp}.xlsx", excel_buf.read())
    zip_buf.seek(0)
    return zip_buf, stamp


# ══════════════════════════════════════════════════════════════════════════════
# Bot Overview Stats
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# Required Channels — CRUD
# ══════════════════════════════════════════════════════════════════════════════

def _get_required_channels() -> list:
    with _lock:
        return list(_U.get("required_channels", []))


def _add_required_channel(name: str, url: str, ch_id: str) -> bool:
    with _lock:
        channels = _U.setdefault("required_channels", [])
        for ch in channels:
            if ch.get("id", "").lower() == ch_id.lower():
                return False
        channels.append({"name": name, "url": url, "id": ch_id})
        _mark_u_dirty()
        return True


def _update_required_channel(index: int, name: str, url: str, ch_id: str) -> bool:
    with _lock:
        channels = _U.get("required_channels", [])
        if index < 0 or index >= len(channels):
            return False
        channels[index] = {"name": name, "url": url, "id": ch_id}
        _mark_u_dirty()
        return True


def _delete_required_channel(index: int) -> bool:
    with _lock:
        channels = _U.get("required_channels", [])
        if index < 0 or index >= len(channels):
            return False
        channels.pop(index)
        _mark_u_dirty()
        return True


def _get_channel_check_interval() -> int:
    with _lock:
        return int(_U["settings"].get("channel_check_interval", "60"))


def _set_channel_check_interval(seconds: int):
    with _lock:
        _U["settings"]["channel_check_interval"] = str(max(10, int(seconds)))
        _mark_u_dirty()


def _get_bot_overview_stats() -> dict:
    """Collect all stats needed for the 📊 Bot Overview screen."""
    with _lock:
        today = _today()

        # ── Users ─────────────────────────────────────────────────────────────
        total_users = len(_U["users"])

        # ── Countries & Numbers ───────────────────────────────────────────────
        counts: dict[int, list] = {}
        for r in _N["numbers"].values():
            cid = int(r["country_id"])
            if cid not in counts:
                counts[cid] = [0, 0]   # [total_added, available]
            counts[cid][0] += 1
            if r["used"] == 0:
                counts[cid][1] += 1

        country_rows = []
        for c in sorted(_N["countries"].values(), key=lambda x: x["name"]):
            cid  = int(c["id"])
            pair = counts.get(cid, [0, 0])
            country_rows.append({
                "name":      c["name"],
                "total":     pair[0],
                "available": pair[1],
                "used":      pair[0] - pair[1],
            })
        total_countries = len(country_rows)

        # ── Panels logged in ──────────────────────────────────────────────────
        panel_statuses = list(_U.get("panel_status", {}).values())
        panels_logged_in = sum(1 for s in panel_statuses if s.get("logged_in"))

        # ── Referrals ─────────────────────────────────────────────────────────
        total_referrals = len(_U.get("referral_log", {}))

    return {
        "total_users":      total_users,
        "total_countries":  total_countries,
        "country_rows":     country_rows,
        "panels_logged_in": panels_logged_in,
        "total_panels":     10,
        "total_referrals":  total_referrals,
    }
